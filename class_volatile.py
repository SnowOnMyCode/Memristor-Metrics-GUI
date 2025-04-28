from PyQt5.QtWidgets import (QApplication, QWidget, QColorDialog, QVBoxLayout, QPushButton, QSlider, QFileDialog, QComboBox, QLabel, QHBoxLayout, QMainWindow, QMenu, QMessageBox, QProgressDialog, QCheckBox, QInputDialog, QLineEdit)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.ticker import LogFormatterMathtext
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas, NavigationToolbar2QT as NavigationToolbar
import mplcursors
import numpy as np
from scipy.signal import find_peaks
from openpyxl import load_workbook
import xlrd
import math

class VolatileSweepsAnalyser(QWidget):
    def __init__(self):
        super().__init__()

        self.init_ui()
        #self.showMaximized()

    def init_ui(self):
        self.setWindowTitle("Volatile Sweeps Viewer")
        #self.setMinimumSize(1200, 800)  # Set a larger size for the window

        # Layout
        main_layout = QVBoxLayout()
        
                
        # Upload button
        self.upload_button = QPushButton("Upload Files")
        self.upload_button.clicked.connect(self.upload_files)
        main_layout.addWidget(self.upload_button)
        self.upload_button.setFixedSize(150, 50)

        # Slider
        self.slider = QSlider(Qt.Horizontal)
        self.slider.setMinimum(1)
        self.slider.valueChanged.connect(self.update_plot)
        main_layout.addWidget(self.slider)

        ##################################################### Plot areas #######################################################################
        main_plot_layout = QHBoxLayout()
        
        self.figure, self.ax1 = plt.subplots(figsize=(10, 6))
        self.canvas = FigureCanvas(self.figure)
        self.canvas.setContextMenuPolicy(Qt.CustomContextMenu)
        self.canvas.customContextMenuRequested.connect(self.show_context_menu)
        main_plot_layout.addWidget(self.canvas, stretch=4)
        main_layout.addLayout(main_plot_layout)

        ######################################################## All Labels ###################################################################
        # Layout for labels to the right of main plot
        main_plot_labels_layout = QVBoxLayout()
        main_plot_labels_layout.setAlignment(Qt.AlignRight)
        main_plot_labels_layout.setSpacing(0)
        main_plot_layout.addLayout(main_plot_labels_layout)
        
        #Making text bigger
        self.setStyleSheet("""
            QLabel {
                font-family: Arial;
                font-size: 16pt;
                color: White;
            }
            QCheckBox {
                font-family: Arial;
                font-size: 12pt;
                color: White;
        """)
        
        #Label for current interval Vset
        self.current_Vset = QLabel("Vset: N/A")
        self.current_Vset.setAlignment(Qt.AlignLeft)
        self.current_Vset.setAlignment(Qt.AlignTop)
        main_plot_labels_layout.addWidget(self.current_Vset)

        # Label for slider counter
        self.slider_counter = QLabel("Cycle nr: N/A")
        self.slider_counter.setAlignment(Qt.AlignLeft)
        self.slider_counter.setAlignment(Qt.AlignTop)
        main_plot_labels_layout.addWidget(self.slider_counter)

        # Label total cycles
        self.total_cycles = QLabel("Total cycles: N/A")
        self.total_cycles.setAlignment(Qt.AlignLeft)
        self.total_cycles.setAlignment(Qt.AlignTop)
        main_plot_labels_layout.addWidget(self.total_cycles)

        # Label Vset
        self.Vset_label = QLabel("Average Vset: N/A")
        self.Vset_label.setAlignment(Qt.AlignLeft)
        self.Vset_label.setAlignment(Qt.AlignTop)
        main_plot_labels_layout.addWidget(self.Vset_label)

        # Color Picker for Voltage Line
        voltage_color_layout = QHBoxLayout()
        voltage_color_label = QLabel("Pick Line Color")
        #voltage_color_layout.addWidget(voltage_color_label)
        self.voltage_color_button = QPushButton("Choose Line Color")
        self.voltage_color_button.setFixedSize(300, 50)
        self.voltage_color_button.clicked.connect(self.pick_voltage_color)
        self.voltage_color_button.adjustSize()
        voltage_color_layout.addWidget(self.voltage_color_button)
        main_plot_labels_layout.addLayout(voltage_color_layout)

        # Color Picker for first Line
        first_color_layout = QHBoxLayout()
        first_color_label = QLabel("Pick first Line Color")
        #voltage_color_layout.addWidget(voltage_color_label)
        self.first_color_button = QPushButton("Choose First Line Color")
        self.first_color_button.setFixedSize(300, 50)
        self.first_color_button.clicked.connect(self.pick_first_color)
        self.first_color_button.adjustSize()
        first_color_layout.addWidget(self.first_color_button)
        main_plot_labels_layout.addLayout(first_color_layout)

        main_plot_labels_layout.addStretch(1)

        # Create a checkbox
        self.checkbox = QCheckBox("Plot all cycles")
        self.checkbox.setChecked(False)  # Initially unchecked
        main_plot_labels_layout.addWidget(self.checkbox, alignment=Qt.AlignBottom)
        self.checkbox.stateChanged.connect(self.update_plot)


        self.setLayout(main_layout)

        ###################################################### Lower Buttons #####################################################
        # Create a horizontal layout for the buttons
        under_graph_hbox = QHBoxLayout()

        # Create a horizontal layout for the buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)  # Set spacing between buttons
        button_layout.setAlignment(Qt.AlignLeft)  # Align buttons to the left

        # Add the button layout to the main layout
        under_graph_hbox.addLayout(button_layout)
        
        main_layout.addLayout(under_graph_hbox)
        
        self.file_paths = ""
        ################## Initialize all to be extracted data ##################
        self.sweeps = []
        self.dfreset = []
        self.dfset = []
        self.Vset = []
        #self.datapts = 300
        self.current_index = 0
        self.voltage_color = 'blue'
        self.first_color = 'deepskyblue'

    def clear_lists(self):
        self.sweeps = []
        self.Vset = []
        self.dfset = []
        self.dfreset = []
        
    ####################### Main funcion of the script, reads excel files, calculates and initializes everything, initializes plots ###########################
    def upload_files(self):
        # Ask for voltage channel name
        VChannel, ok1 = QInputDialog.getText(self, "Voltage Channel", "Enter the voltage channel name:", QLineEdit.Normal, "DrainV")
        if not ok1:
            return
        elif not VChannel:
            VChannel = "DrainV"  # Default value if canceled or empty

        # Ask for current channel name
        IChannel, ok2 = QInputDialog.getText(self, "Current Channel", "Enter the current channel name:", QLineEdit.Normal, "DrainI")
        if not ok2:
            return
        elif not IChannel:
            IChannel = "DrainI"  # Default value if canceled or empty
        
        # Assign the user-input column names
        self.VChannel = VChannel
        self.IChannel = IChannel

        try:
            options = QFileDialog.Options()
            self.file_paths, _ = QFileDialog.getOpenFileNames(self, "Upload Excel Files", "", "Excel Files (*.xls *.xlsx)", options=options)
            
            if len(self.file_paths) == 0:   ## Do nothing if there are no selected files i.e. cancel was clicked
                return
            if len(self.sweeps) != 0:   ## If other files are uploaded again empty the existing sweeps
                self.clear_lists()

            # Initialize progress dialog
            progress = QProgressDialog("Loading files...", "Cancel", 0, len(self.file_paths), self)
            progress.setWindowModality(Qt.WindowModal)
            progress.setAutoClose(True)
            progress.setAutoReset(True)

            # Read all sheets in all selected files except for the "Calc" and "Settings" sheets.
            total_steps = 0
            for file_path in self.file_paths:
                if file_path.endswith('.xlsx'):
                    total_steps += len(load_workbook(file_path, read_only=True).sheetnames) - 2  # Calculate total steps
                else:
                    total_steps += len(xlrd.open_workbook(file_path).sheets()) - 2
            progress.setRange(0, total_steps)
            current_step = 0

            ## Read all sheets in all selected files except for the "Calc" and "Settings" sheets.
            for file_path in self.file_paths:
                if progress.wasCanceled():
                    return  # Handle cancellation

                excel_file = pd.ExcelFile(file_path)
                sheet_names = [sheet for sheet in excel_file.sheet_names if sheet not in ["Calc", "Settings"]]
                
                for sheet in sheet_names:
                    if progress.wasCanceled():
                        return  # Handle cancellation
                    df = pd.read_excel(file_path, sheet_name=sheet)
                    if df.empty:    #Ignore empty sheets
                        progress.setValue(current_step)
                        current_step += 1
                        continue
                    self.sweeps.append(df)
                    current_step += 1
                    progress.setValue(current_step)

            # Filter the sweeps to remove those where the first value of the Voltage column is NaN
            #self.sweeps = [sweep for sweep in self.sweeps if not pd.isna(sweep.DrainV.iloc[0])]

            #if len(self.sweeps) == 0:
                #raise KeyError("Excel files provided have no data")
                        
            self.slider.setMaximum(len(self.sweeps))
            self.current_index = 0

            self.slider.setValue(0)
            self.update_plot()

            self.total_cycles.setText(f"Total cycles: {len(self.sweeps)} ")
        except KeyError as e:
            QMessageBox.critical(self, "Error", f"Voltage column doesn't have data: {str(e)}. Please try another file.")

    def update_plot(self, *args):
        self.current_index = self.slider.value()-1
        if self.file_paths == "":
            return
        elif len(self.sweeps) == 0:
            return
        # Update the dataframe with the step of the slider
        if self.checkbox.isChecked():
            self.slider.setEnabled(False)  # Disable the slider
            # Clear the previous plot
            self.figure.clear()

            # Create new axes
            self.ax1 = self.figure.add_subplot(111)

            # Left axis for Voltage Channel
            self.ax1.set_xlabel(r'$\it{V}\ (V)$')
            self.ax1.set_ylabel(r'$\it{I}\ (A)$')
            for i,df in enumerate(reversed(self.sweeps)):
                if i == len(self.sweeps) - 1:
                    line1, = self.ax1.plot(df[self.VChannel], df[self.IChannel].abs(), color=self.first_color, linewidth=2, alpha=0.8, label=self.VChannel)
                else:
                    line1, = self.ax1.plot(df[self.VChannel], df[self.IChannel].abs(), color=self.voltage_color, linewidth=2, alpha=0.8, label=self.VChannel)
            #self.ax1.plot(df[self.VChannel], df[self.IChannel].abs(), color="blue", linewidth=0.5)
            self.ax1.tick_params(axis='y', labelcolor='black')
            #self.ax1.grid(True, which='both', linestyle='--', linewidth=0.5)
            self.ax1.set_yscale('log')

            # Format y-axis with scientific notation
            self.ax1.yaxis.set_major_formatter(LogFormatterMathtext(base=10))

            # Add interactive cursor
            mplcursors.cursor([line1], hover=2)
            
            self.figure.tight_layout()
            self.canvas.draw()
        else:
            self.slider.setEnabled(True)  # Enable the slider
            df = self.sweeps[self.current_index]
            self.plot_data(df)

            # Update Vset and Vreset labels
            self.slider_counter.setText(f"Cycle nr: {self.current_index+1} ")
            temp_df = self.sweeps[self.current_index]
            
            self.current_Vset.setText(f"Vset: {self.write_Vset(temp_df):.3f} ")
        
    
    def write_Vset(self, df):
        forth_idxes = np.where(np.gradient(df[self.VChannel])>0)[0]
        forth_idxes = forth_idxes[df[self.VChannel].iloc[forth_idxes] > 0.05]
        setvoltage = df[self.VChannel].iloc[forth_idxes]
        setcurrent = df[self.IChannel].iloc[forth_idxes]
        settime = df.Time.iloc[forth_idxes]
        R = setvoltage.abs()/setcurrent.abs()
        dR = -(np.gradient(np.log(R)))
        max_idx,_ = find_peaks(dR)
        
        highest_peak = max_idx[np.argmax(dR[max_idx])]
        threshold = 0.7 * dR[highest_peak]
        max_idx2 = max_idx[dR[max_idx] >= threshold]
        max_dR = dR[max_idx2]
        max_t = settime.iloc[max_idx2]

        if max_dR.size == 0:
            VSET = np.min(setvoltage)
        else:
            idx_RESET = np.where(settime == max_t.iloc[-1])[0][0]
            VSET = setvoltage.iloc[idx_RESET]
            if abs(VSET) < 0.05:
                VSET = setvoltage.iloc[-1]
        return VSET
    
    def update_Vset(self, dfset):
        if len(dfset) == 0:
            return
        for set_sweep in dfset:
            # Take only the sweep in and not the sweep back part 
            # and also get rid of the first part of the sweep because it messes with the gradient
            forth_idxes = np.where(np.gradient(set_sweep.Voltage)>0)[0]
            forth_idxes = forth_idxes[set_sweep.Voltage.iloc[forth_idxes] > 0.05]
            setvoltage = set_sweep.Voltage.iloc[forth_idxes]
            setcurrent = set_sweep.Current.iloc[forth_idxes]
            settime = set_sweep.TimeOutput.iloc[forth_idxes]
            R = setvoltage.abs()/setcurrent.abs()
            dR = -(np.gradient(np.log(R)))
            max_idx,_ = find_peaks(dR)
            
            highest_peak = max_idx[np.argmax(dR[max_idx])]
            threshold = 0.7 * dR[highest_peak]
            max_idx2 = max_idx[dR[max_idx] >= threshold]
            max_dR = dR[max_idx2]
            max_t = settime.iloc[max_idx2]

            if max_dR.size == 0:
                VSET = np.min(setvoltage)
            else:
                idx_RESET = np.where(settime == max_t.iloc[-1])[0][0]
                VSET = setvoltage.iloc[idx_RESET]
                if abs(VSET) < 0.05:
                    VSET = setvoltage.iloc[-1]
            self.Vset.append(VSET)

    def plot_data(self, df):
        # Clear the previous plot
        self.figure.clear()

        # Create new axes
        self.ax1 = self.figure.add_subplot(111)

        # Left axis for Voltage Channel
        self.ax1.set_xlabel(r'$\it{V}\ (V)$')
        self.ax1.set_ylabel(r'$\it{I}\ (A)$')
        line1, = self.ax1.plot(df[self.VChannel], df[self.IChannel].abs(), color=self.voltage_color, linewidth=2, alpha=0.8, label=self.VChannel)
        self.ax1.tick_params(axis='y', labelcolor='black')
        self.ax1.set_yscale('log')

        # Format y-axis with scientific notation
        self.ax1.yaxis.set_major_formatter(LogFormatterMathtext(base=10))

        # Add interactive cursor
        mplcursors.cursor([line1], hover=2)
        
        self.figure.tight_layout()
        self.canvas.draw()

    def show_context_menu(self, pos):
        context_menu = QMenu(self)
        save_action = context_menu.addAction("Save Plot")
        action = context_menu.exec_(self.canvas.mapToGlobal(pos))
        if action == save_action:
            self.save_plot(self.figure)

    def pick_voltage_color(self):
        # Open color picker dialog and set the selected color
        color = QColorDialog.getColor()
        if color.isValid():
            self.voltage_color = color.name()  # Convert selected color to HEX code (e.g., "#RRGGBB")
            self.voltage_color_button.setStyleSheet(f"background-color: {self.voltage_color};")  # Update button background color
        self.update_plot()

    def pick_first_color(self):
        # Open color picker dialog and set the selected color
        color = QColorDialog.getColor()
        if color.isValid():
            self.first_color = color.name()  # Convert selected color to HEX code (e.g., "#RRGGBB")
            self.first_color_button.setStyleSheet(f"background-color: {self.first_color};")  # Update button background color
        self.update_plot()