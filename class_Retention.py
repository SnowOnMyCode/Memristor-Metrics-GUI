import sys
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, QPushButton, QSlider, QFileDialog, QComboBox, QLabel, QHBoxLayout, QMainWindow, QMessageBox, QProgressBar, QProgressDialog,QDialog)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.ticker import FuncFormatter, AutoLocator, LogFormatterMathtext
import mplcursors
from scipy.signal import find_peaks
from openpyxl import load_workbook

class Retention_viewer(QWidget):
    def __init__(self):
        super().__init__()

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Retention Viewer")

        # Layout
        main_layout = QVBoxLayout()

        # Upload LRS button
        self.upload_button_LRS = QPushButton("Upload LRS Excel Files")
        self.upload_button_LRS.clicked.connect(self.upload_files_LRS)
        main_layout.addWidget(self.upload_button_LRS)

        # Upload HRS button
        self.upload_button_HRS = QPushButton("Upload HRS Excel Files")
        self.upload_button_HRS.clicked.connect(self.upload_files_HRS)
        main_layout.addWidget(self.upload_button_HRS)

        # Slider LRS
        self.slider_LRS = QSlider(Qt.Horizontal)
        self.slider_LRS.setMinimum(1)
        self.slider_LRS.valueChanged.connect(self.plot_data)
        main_layout.addWidget(self.slider_LRS)

        # Slider HRS
        self.slider_HRS = QSlider(Qt.Horizontal)
        self.slider_HRS.setMinimum(1)
        self.slider_HRS.valueChanged.connect(self.plot_data)
        main_layout.addWidget(self.slider_HRS)

        # Plot area
        self.figure = plt.figure(figsize=(14,6))
        self.ax1 = self.figure.add_subplot(111)
        self.canvas = FigureCanvas(self.figure)
        main_layout.addWidget(self.canvas)

        # Clear Graph button
        self.clear_button = QPushButton("Clear Graph")
        self.clear_button.clicked.connect(self.clear_graph)
        main_layout.addWidget(self.clear_button)

        self.setLayout(main_layout)

        self.current_index = 0
        self.RChannel = "R"  # Default current channel
        self.TimeChannel = "Time" # Default time channel

        self.sweeps = []
        self.Time_LRS = []
        self.R_LRS= []

        self.Time_HRS = []
        self.R_HRS = []

        # Label to display HRS Avg
        self.HRS_label = QLabel("HRS average: N/A")
        self.HRS_label.setAlignment(Qt.AlignRight)
        main_layout.addWidget(self.HRS_label)

        # Label to display HRS Avg
        self.LRS_label = QLabel("LRS average: N/A")
        self.LRS_label.setAlignment(Qt.AlignRight)
        main_layout.addWidget(self.LRS_label)

    def clear_graph(self):
        self.sweeps = []
        self.Time_LRS = []
        self.R_LRS= []
        self.Time_HRS = []
        self.R_HRS = []
        self.HRS_label.setText(f"HRS average: N/A")
        self.LRS_label.setText(f"LRS average: N/A")
        self.plot_data()


    def upload_files_HRS(self):
        try:
            options = QFileDialog.Options()
            self.file_paths, _ = QFileDialog.getOpenFileNames(self, "Upload Excel Files", "", "Excel Files (*.xls *.xlsx)", options=options)
            
            if len(self.file_paths) == 0:   ## Do nothing if there are no selected files i.e. cancel was clicked
                return
            if len(self.R_HRS) != 0:   ## If other files are uploaded again empty the existing sweeps
                self.R_HRS= []
                self.Time_HRS = []
                self.sweeps = []

            # Initialize progress dialog
            progress = QProgressDialog("Loading files...", "Cancel", 0, len(self.file_paths), self)
            progress.setWindowModality(Qt.WindowModal)
            progress.setAutoClose(True)
            progress.setAutoReset(True)

            # Calculations needed for progress bar
            total_steps = 0
            for file_path in self.file_paths:
                if file_path.endswith('.xlsx'):
                    total_steps += len(load_workbook(file_path, read_only=True).sheetnames) - 2  # Calculate total steps
                else:
                    xl = pd.ExcelFile(file_path, engine = 'xlrd')
                    total_steps += len(xl.sheet_names) - 2
            progress.setRange(0, total_steps)
            current_step = 0
            last_time = 0
            # Read all sheets in all selected files except for the "Calc" and "Settings" sheets.
            for file_path in self.file_paths:
                if progress.wasCanceled():
                    return  # Handle cancellation

                excel_file = pd.ExcelFile(file_path)
                sheet_names = [sheet for sheet in excel_file.sheet_names if sheet not in ["Calc", "Settings"]]

                for sheet in sheet_names:
                    if progress.wasCanceled():
                        return  # Handle cancellation
                    sweep = pd.read_excel(file_path, sheet_name=sheet)
                    #self.sweeps.append(pd.read_excel(file_path, sheet_name=sheet))
                    if self.TimeChannel not in sweep or self.RChannel not in sweep:
                        raise KeyError(f"Sheet '{sheet}' in file '{file_path}' does not contain the required columns.")
                    
                    # Append the data after adjusting the time
                    self.R_HRS.append(sweep[self.RChannel])  # Append the 'R' data
                    
                    # Adjust the 'Time' column by adding the last_time
                    self.Time_HRS.append(sweep[self.TimeChannel] + last_time)

                    # Update last_time to the last value of the current sheet's 'Time'
                    last_time = sweep[self.TimeChannel].iloc[-1]  # Accumulate the time

                    current_step += 1
                    progress.setValue(current_step)
            
            # After processing all sheets, concatenate the lists into a single pandas Series
            self.R_HRS= pd.concat(self.R_HRS, ignore_index=True)
            self.Time_HRS = pd.concat(self.Time_HRS, ignore_index=True)

            self.slider_HRS.setMaximum(len(self.R_HRS)-1)
            self.current_index = 0

            self.slider_HRS.setValue(0)
            self.plot_data()

            self.HRS_label.setText(f"HRS average: {np.mean(self.R_HRS):.3e}")

            self.plot_data()

        except KeyError as e:
            QMessageBox.critical(self, "Error", f"Voltage column doesn't have data: {str(e)}. Please try another file.")

    def upload_files_LRS(self):
        try:
            options = QFileDialog.Options()
            self.file_paths, _ = QFileDialog.getOpenFileNames(self, "Upload Excel Files", "", "Excel Files (*.xls *.xlsx)", options=options)
            
            if len(self.file_paths) == 0:   ## Do nothing if there are no selected files i.e. cancel was clicked
                return
            if len(self.R_LRS) != 0:   ## If other files are uploaded again empty the existing sweeps
                self.R_LRS= []
                self.Time_LRS = []
                self.sweeps = []

            # Initialize progress dialog
            progress = QProgressDialog("Loading files...", "Cancel", 0, len(self.file_paths), self)
            progress.setWindowModality(Qt.WindowModal)
            progress.setAutoClose(True)
            progress.setAutoReset(True)

            # Calculations needed for progress bar
            total_steps = 0
            for file_path in self.file_paths:
                if file_path.endswith('.xlsx'):
                    total_steps += len(load_workbook(file_path, read_only=True).sheetnames) - 2  # Calculate total steps
                else:
                    xl = pd.ExcelFile(file_path, engine = 'xlrd')
                    total_steps += len(xl.sheet_names) - 2
            progress.setRange(0, total_steps)
            current_step = 0

            # Read all sheets in all selected files except for the "Calc" and "Settings" sheets.
            for file_path in self.file_paths:
                if progress.wasCanceled():
                    return  # Handle cancellation

                excel_file = pd.ExcelFile(file_path)
                sheet_names = [sheet for sheet in excel_file.sheet_names if sheet not in ["Calc", "Settings"]]
                
                last_time = 0

                for sheet in sheet_names:
                    if progress.wasCanceled():
                        return  # Handle cancellation
                    sweep = pd.read_excel(file_path, sheet_name=sheet)
                    #self.sweeps.append(pd.read_excel(file_path, sheet_name=sheet))
                    if self.TimeChannel not in sweep or self.RChannel not in sweep:
                        raise KeyError(f"Sheet '{sheet}' in file '{file_path}' does not contain the required columns.")
                    
                    # Append the data after adjusting the time
                    self.R_LRS.append(sweep[self.RChannel])  # Append the 'R' data
                    
                    # Adjust the 'Time' column by adding the last_time
                    self.Time_LRS.append(sweep[self.TimeChannel] + last_time)

                    # Update last_time to the last value of the current sheet's 'Time'
                    last_time = sweep[self.TimeChannel].iloc[-1]  # Accumulate the time

                    current_step += 1
                    progress.setValue(current_step)
            
            # After processing all sheets, concatenate the lists into a single pandas Series
            self.R_LRS= pd.concat(self.R_LRS, ignore_index=True)
            self.Time_LRS = pd.concat(self.Time_LRS, ignore_index=True)

            self.slider_LRS.setMaximum(len(self.R_LRS)-1)
            self.current_index = 0

            self.slider_LRS.setValue(0)
            self.plot_data()

            self.LRS_label.setText(f"LRS average: {np.mean(self.R_LRS):.3e}")

            self.plot_data()

        except KeyError as e:
            QMessageBox.critical(self, "Error", f"Voltage column doesn't have data: {str(e)}. Please try another file.")

    def plot_data(self):
        try:
            self.current_index_LRS = self.slider_LRS.value()-1
            self.current_index_HRS = self.slider_HRS.value()-1
            # Clear the previous plot
            self.ax1.clear()

            # Left axis for Voltage Channel
            self.ax1.set_xlabel('Time (s)')
            self.ax1.set_xscale('log')
            self.ax1.set_yscale('log')
            self.ax1.set_ylabel('Resistance (Ω)')
            line1, = self.ax1.plot(self.Time_LRS[0:self.current_index_LRS], self.R_LRS[0:self.current_index_LRS], color="red",linestyle='none', marker='o', label='LRS')
            if len(self.R_HRS) != 0:
                line2, = self.ax1.plot(self.Time_HRS[0:self.current_index_HRS], self.R_HRS[0:self.current_index_HRS], color="blue",linestyle='none', marker= 'o', label="HRS")
                mplcursors.cursor([line2], hover=2)
            self.ax1.tick_params(axis='y', labelcolor='black')
            #self.ax1.grid(True, which='both', linestyle='--', linewidth=0.5)
            self.ax1.legend()
            # Format y-axis with scientific notation
            self.ax1.yaxis.set_major_formatter(LogFormatterMathtext(base=10))

            # Add interactive cursor
            mplcursors.cursor([line1], hover=2)

            self.figure.tight_layout()
            self.canvas.draw()

        except KeyError as e:
            QMessageBox.critical(self, "Error", f"Column doesn't exist: {str(e)}. Please try another file.")