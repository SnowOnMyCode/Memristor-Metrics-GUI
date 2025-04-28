from PyQt5.QtWidgets import (QApplication, QColorDialog, QWidget, QVBoxLayout, QPushButton, QSlider, QFileDialog, QComboBox, QLabel, QHBoxLayout, QMainWindow, QMenu, QMessageBox, QProgressDialog, QCheckBox)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas, NavigationToolbar2QT as NavigationToolbar
from matplotlib.ticker import LogFormatterMathtext
import mplcursors
import numpy as np
from scipy.signal import find_peaks
from openpyxl import load_workbook
import xlrd
import math

class IVSweepsAnalyser(QWidget):
    def __init__(self):
        super().__init__()

        self.init_ui()
        self.showMaximized()

    def init_ui(self):
        self.setWindowTitle("IV Sweeps Viewer") 
        self.setMinimumSize(1200, 800)  # Set a larger size for the window

        # Layout
        main_layout = QVBoxLayout()

        # Slider
        self.slider = QSlider(Qt.Horizontal)
        self.slider.setMinimum(1)
        self.slider.valueChanged.connect(self.update_plot)
        main_layout.addWidget(self.slider)

        ##################################################### Plot areas #######################################################################
        main_plot_layout = QHBoxLayout()
        
        self.figure, self.ax1 = plt.subplots(figsize=(14, 6))
        self.canvas = FigureCanvas(self.figure)
        self.canvas.setContextMenuPolicy(Qt.CustomContextMenu)
        self.canvas.customContextMenuRequested.connect(self.show_context_menu)
        main_plot_layout.addWidget(self.canvas, stretch=2)
        main_layout.addLayout(main_plot_layout)

        # Additional plot areas
        self.plots_layout = QHBoxLayout()
        
        # Dropdown menu to select plot type
        self.plot_type_dropdown = QComboBox()
        self.plot_type_dropdown.addItems(["CDF Plots", "Resistance Plot"])
        self.plot_type_dropdown.currentIndexChanged.connect(self.update_plot_visibility)
        main_layout.addWidget(self.plot_type_dropdown)


        self.figure2, self.ax2 = plt.subplots(figsize=(7, 6))
        self.canvas2 = FigureCanvas(self.figure2)
        self.canvas2.setContextMenuPolicy(Qt.CustomContextMenu)
        self.canvas2.customContextMenuRequested.connect(self.show_context_menu2)
        self.plots_layout.addWidget(self.canvas2)
        
        
        self.figure3, self.ax3 = plt.subplots(figsize=(7, 6))
        self.canvas3 = FigureCanvas(self.figure3)
        self.canvas3.setContextMenuPolicy(Qt.CustomContextMenu)
        self.canvas3.customContextMenuRequested.connect(self.show_context_menu3)
        self.plots_layout.addWidget(self.canvas3)

        self.figure4, self.ax4 = plt.subplots(figsize=(14, 6))
        self.canvas4 = FigureCanvas(self.figure4)
        self.canvas4.setContextMenuPolicy(Qt.CustomContextMenu)
        self.canvas4.customContextMenuRequested.connect(self.show_context_menu4)
        self.plots_layout.addWidget(self.canvas4)
        self.canvas4.setVisible(False)

        self.canvas2.setMinimumSize(400,400)
        self.canvas3.setMinimumSize(400,400)
        self.canvas4.setMinimumSize(400,400)
        main_layout.addLayout(self.plots_layout)

        ######################################################## All Labels ###################################################################
        # Layout for labels to the right of main plot
        main_plot_labels_layout = QVBoxLayout()
        main_plot_labels_layout.setAlignment(Qt.AlignRight)
        main_plot_labels_layout.setSpacing(0)
        main_plot_layout.addLayout(main_plot_labels_layout)
        
        #Making text bigger
        self.setStyleSheet("""
            QLabel {
                font-family: Arial;
                font-size: 16pt;
                color: White;
            }
            QCheckBox {
                font-family: Arial;
                font-size: 12pt;
                color: White;
        """)

        # Label total cycles
        self.total_cycles = QLabel("Total cycles: N/A")
        self.total_cycles.setAlignment(Qt.AlignLeft)
        self.total_cycles.setAlignment(Qt.AlignTop)
        main_plot_labels_layout.addWidget(self.total_cycles)

        # Label for slider counter
        self.slider_counter = QLabel("Halfcycle nr: N/A")
        self.slider_counter.setAlignment(Qt.AlignLeft)
        self.slider_counter.setAlignment(Qt.AlignTop)
        main_plot_labels_layout.addWidget(self.slider_counter)

        #Label for current interval Vset
        self.current_Vset = QLabel("Vset: N/A")
        self.current_Vset.setAlignment(Qt.AlignLeft)
        self.current_Vset.setAlignment(Qt.AlignTop)
        main_plot_labels_layout.addWidget(self.current_Vset)

        #Label for current interval Vreset
        self.current_Vreset = QLabel("Vreset: N/A")
        self.current_Vreset.setAlignment(Qt.AlignLeft)
        self.current_Vreset.setAlignment(Qt.AlignTop)
        main_plot_labels_layout.addWidget(self.current_Vreset)

        # Label Vreset
        self.Vreset_label = QLabel("Average Vreset: N/A")
        self.Vreset_label.setAlignment(Qt.AlignLeft)
        self.Vreset_label.setAlignment(Qt.AlignTop)
        main_plot_labels_layout.addWidget(self.Vreset_label)

        # Label Vset
        self.Vset_label = QLabel("Average Vset: N/A")
        self.Vset_label.setAlignment(Qt.AlignLeft)
        self.Vset_label.setAlignment(Qt.AlignTop)
        main_plot_labels_layout.addWidget(self.Vset_label)

        # Color Picker for Voltage Line
        voltage_color_layout = QHBoxLayout()
        voltage_color_label = QLabel("Pick Voltage Line Color")
        #voltage_color_layout.addWidget(voltage_color_label)
        self.voltage_color_button = QPushButton("Choose Voltage Color")
        self.voltage_color_button.setFixedSize(300, 50)
        self.voltage_color_button.clicked.connect(self.pick_voltage_color)
        self.voltage_color_button.adjustSize()
        voltage_color_layout.addWidget(self.voltage_color_button)
        main_plot_labels_layout.addLayout(voltage_color_layout)

        main_plot_labels_layout.addStretch(1)

        # Create a checkbox for plotting all cycles
        self.checkbox = QCheckBox("Plot only current interval")
        self.checkbox.setChecked(False)  # Initially unchecked
        main_plot_labels_layout.addWidget(self.checkbox, alignment=Qt.AlignBottom)
        self.checkbox.stateChanged.connect(self.update_plot)

        # Create a checkbox for grid on/off
        self.checkbox_grid = QCheckBox("Grid on/off")
        self.checkbox_grid.setChecked(False)  # Initially unchecked
        main_plot_labels_layout.addWidget(self.checkbox_grid, alignment=Qt.AlignBottom)
        self.checkbox_grid.stateChanged.connect(self.update_plot)

        # Create a checkbox for ignoring bad sheets (missing sets/resets)
        self.checkbox_ignore = QCheckBox("Ignore faulty sheets")
        self.checkbox_ignore.setChecked(True)  # Initially unchecked
        main_plot_labels_layout.addWidget(self.checkbox_ignore, alignment=Qt.AlignBottom)
        #self.checkbox_ignore.stateChanged.connect(self.update_plot)

        self.setLayout(main_layout)

        ###################################################### Lower Buttons #####################################################
        # Create a horizontal layout for the buttons
        under_graph_hbox = QHBoxLayout()

        # Create a horizontal layout for the buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)  # Set spacing between buttons
        button_layout.setAlignment(Qt.AlignRight)  # Align buttons to the left

        # Color Picker for Set Line
        set_color_layout = QHBoxLayout()
        set_color_label = QLabel("Pick Set Line Color")
        self.set_color_button = QPushButton("Choose Set Color")
        self.set_color_button.setFixedSize(300, 50)
        self.set_color_button.clicked.connect(self.pick_set_color)
        self.set_color_button.adjustSize()
        set_color_layout.addWidget(self.set_color_button, stretch=2)
        under_graph_hbox.addLayout(set_color_layout)

        #color picker for reset line
        reset_color_label = QLabel("Pick Reset Line Color")
        self.reset_color_button = QPushButton("Choose Reset Color")
        self.reset_color_button.setFixedSize(300, 50)
        self.reset_color_button.clicked.connect(self.pick_reset_color)
        self.reset_color_button.adjustSize()
        set_color_layout.addWidget(self.reset_color_button, stretch=2)
        under_graph_hbox.addLayout(set_color_layout)

        # Upload button
        self.upload_button = QPushButton("Upload Files")
        self.upload_button.clicked.connect(self.upload_files)
        button_layout.addWidget(self.upload_button)
        self.upload_button.setFixedSize(150, 50)

        # Export Vreset button
        self.export_button = QPushButton("Save Data")
        self.export_button.clicked.connect(self.export)
        button_layout.addWidget(self.export_button)
        self.export_button.setFixedSize(150, 50)

        # Add the button layout to the main layout
        under_graph_hbox.addLayout(button_layout)
        
        main_layout.addLayout(under_graph_hbox)
        
        self.file_paths = ""
        ################## Initialize all to be extracted data ##################
        self.sweeps = []
        self.dfreset = []
        self.dfset = []
        self.Vreset = []
        self.Vset = []
        self.Rreset = []
        self.Rset = []
        self.ResVoltage = []
        self.SetVoltage = []
        self.datapts = 300
        self.current_index = 0
        self.VChannel = "Voltage"  # Default voltage channel
        self.IChannel = "Current"  # Default current channel
        self.set_color = 'blue' # Default set color
        self.reset_color = 'red' # Default reset color
        self.voltage_color = 'red'  # Default voltage color

    ####################### Switch between CDF plots and Resistance plot#################################
    def update_plot_visibility(self):
        selected_plot = self.plot_type_dropdown.currentText()
        if selected_plot == "CDF Plots":
            self.canvas2.setVisible(True)
            self.canvas3.setVisible(True)
            self.canvas4.setVisible(False)
        elif selected_plot == "Resistance Plot":
            self.canvas2.setVisible(False)
            self.canvas3.setVisible(False)
            self.canvas4.setVisible(True)
        # Ensure layout updates are applied
        self.plots_layout.invalidate()
        self.plots_layout.update()

    ####################### Function used in upload_files to clear everything ###########################
    def clear_lists(self):
        self.sweeps = []
        self.Rset = []
        self.Rreset = []
        self.Vset = []
        self.Vreset = []
        self.dfset = []
        self.dfreset = []
        self.ResVoltage = []
        self.SetVoltage = []
        
    ####################### Main funcion of the script, reads excel files, calculates and initializes everything, initializes plots ###########################
    def upload_files(self):
        try:
            options = QFileDialog.Options()
            self.file_paths, _ = QFileDialog.getOpenFileNames(self, "Upload Excel Files", "", "Excel Files (*.xls *.xlsx)", options=options)
            
            if len(self.file_paths) == 0:   ## Do nothing if there are no selected files i.e. cancel was clicked
                return
            if len(self.sweeps) != 0:   ## If other files are uploaded again empty the existing sweeps
                self.clear_lists()

            # Initialize progress dialog
            progress = QProgressDialog("Loading files...", "Cancel", 0, len(self.file_paths), self)
            progress.setWindowModality(Qt.WindowModal)
            progress.setAutoClose(True)
            progress.setAutoReset(True)

            # Read all sheets in all selected files except for the "Calc" and "Settings" sheets.
            total_steps = 0
            for file_path in self.file_paths:
                if file_path.endswith('.xlsx'):
                    total_steps += len(load_workbook(file_path, read_only=True).sheetnames) - 2  # Calculate total steps
                else:
                    total_steps += len(xlrd.open_workbook(file_path).sheets()) - 2
            progress.setRange(0, total_steps)
            current_step = 0

            ## Read all sheets in all selected files except for the "Calc" and "Settings" sheets.
            for file_path in self.file_paths:
                if progress.wasCanceled():
                    return  # Handle cancellation

                excel_file = pd.ExcelFile(file_path)
                sheet_names = [sheet for sheet in excel_file.sheet_names if sheet not in ["Calc", "Settings"]]
                
                for sheet in sheet_names:
                    if progress.wasCanceled():
                        return  # Handle cancellation
                    self.sweeps.append(pd.read_excel(file_path, sheet_name=sheet))
                    current_step += 1
                    progress.setValue(current_step)
                    
            # Filter the sweeps to remove those where the first value of the Voltage column is NaN
            self.sweeps = [sweep for sweep in self.sweeps if not pd.isna(sweep.Voltage.iloc[0])]
            if self.checkbox_ignore.isChecked():
                self.sweeps = [sweep for sweep in self.sweeps if not sweep.ResetVoltage.notna().sum()!=sweep.SetVoltage.notna().sum()]

            if len(self.sweeps) == 0:
                raise KeyError("Excel files provided have no data")

            ## Extract Rset, Rreset and zero indexes from an array 
            for sweep in self.sweeps:
                self.Rset.append(sweep.SetResistance)
                self.Rreset.append(sweep.ResetResistance)
                self.ResVoltage.append(sweep.ResetVoltage)
                self.SetVoltage.append(sweep.SetVoltage)


            self.sweeps = pd.concat(self.sweeps, ignore_index=True)
            self.Rset = pd.concat(self.Rset, ignore_index=True)
            self.Rset = self.Rset[~np.isnan(self.Rset)]
            self.Rset = self.Rset.to_numpy()
            self.Rreset = pd.concat(self.Rreset, ignore_index=True)
            self.Rreset = self.Rreset[~np.isnan(self.Rreset)]
            self.ResVoltage = pd.concat(self.ResVoltage, ignore_index=True)
            self.ResVoltage = self.ResVoltage[~np.isnan(self.ResVoltage)]
            self.SetVoltage = pd.concat(self.SetVoltage, ignore_index=True)
            self.SetVoltage = self.SetVoltage[~np.isnan(self.SetVoltage)]

            no_reset_idx = np.where(self.ResVoltage==0)[0].tolist()
            no_set_idx = np.where(self.SetVoltage==0)[0].tolist()
            
            # Update zero indexes and datapoints (usually 300)
            zero_idxes = [index for index, value in enumerate(self.sweeps.Voltage) if value == 0]
            
            self.datapts = zero_idxes[1]
            zero_idxes.append(zero_idxes[-1]+(self.datapts-1))
        
            for i in zero_idxes[1:]:
                df = self.sweeps[(i-self.datapts):i]
                if df.Voltage[i+2-self.datapts] < 0:
                    self.dfreset.append(df)
                else:
                    self.dfset.append(df)
            
            self.slider.setMaximum(len(zero_idxes)-1)
            self.current_index = 0

            self.slider.setValue(0)
            self.update_plot()

            #self.sweeps = [sweep for sweep in self.sweeps if not sweep.ResetVoltage.notna().sum()!=sweep.SetVoltage.notna().sum()]

            self.update_Vreset(self.dfreset)
            self.update_Vset(self.dfset)

            ## Where Set or Reset didn't happen delete Vset/Vreset and Rset/Rreset
            for idx in sorted(no_set_idx, reverse=False):
                self.Vset.insert(idx, math.nan)
                self.Rset = np.insert(self.Rset, idx, np.nan)

            for idx in sorted(no_reset_idx,reverse=False):
                self.Vreset.insert(idx, math.nan)
                self.Rreset = np.insert(self.Rreset, idx, np.nan)


            for index in sorted(no_set_idx,reverse=True):
                del self.Vreset[index]
            self.Rreset = np.delete(self.Rreset,no_set_idx)
            for index in sorted(no_reset_idx,reverse=True):
                del self.Vset[index]
            self.Rset = np.delete(self.Rset,no_reset_idx)
            
            #delete all nans 
            self.Vset = [x for x in self.Vset if not math.isnan(x)]
            self.Vreset = [x for x in self.Vreset if not math.isnan(x)]
            self.Rset = self.Rset[~np.isnan(self.Rset)]
            self.Rreset = self.Rreset[~np.isnan(self.Rreset)]

            self.Vreset_label.setText(f"Average Vreset: {np.mean(self.Vreset):.3f} ")
            self.Vset_label.setText(f"Average Vset: {np.mean(self.Vset):.3f} ")
            self.total_cycles.setText(f"Total cycles: {(len(zero_idxes)-1)/2} ")
        except KeyError as e:
            QMessageBox.critical(self, "Error", f"Voltage column doesn't have data: {str(e)}. Please try another file.")

    ########################### Update-ing plot after slider move or checkbox change #################################
    def update_plot(self, *args):
        self.current_index = self.slider.value()-1
        if self.file_paths == "":
            return
        elif len(self.sweeps) == 0:
            return
        # Update the dataframe with the step of the slider
        if self.checkbox.isChecked():
            df = self.sweeps[self.datapts*(self.current_index):self.datapts*(self.current_index+1)]
            self.plot_data(df)
        else:
            df = self.sweeps[0:self.datapts*(self.current_index+1)]
            self.plot_data(df)

        self.plot_cdf_V()
        self.plot_cdf_R()
        self.plot_res()

        # Update Vset and Vreset labels
        self.slider_counter.setText(f"Halfcycle nr: {self.current_index+1} ")
        temp_df = self.sweeps[self.datapts*(self.current_index):self.datapts*(self.current_index+1)]
        if temp_df.Voltage[self.datapts*(self.current_index)+1] > 0:
            self.current_Vset.setText(f"Vset: {self.write_Vset(temp_df):.3f} ")
            self.current_Vreset.setText("Vreset: N/A")
        else:
            self.current_Vreset.setText(f"Vreset: {self.write_Vreset(temp_df):.3f} ")
            self.current_Vset.setText("Vset: N/A")

    ############################ Save Data ##########################################
    def export(self):
        
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Data to Excel", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if file_path:
            # Ensure the file has the correct extension
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'

            # Make sure Vreset and Vset are the same length
            max_length = max(len(self.Vreset), len(self.Vset), len(self.Rreset), len(self.Rset))
            Vreset_padded = np.pad(self.Vreset, (0, max_length - len(self.Vreset)), constant_values=np.nan)
            Vset_padded = np.pad(self.Vset, (0, max_length - len(self.Vset)), constant_values=np.nan)
            Rset_padded = np.pad(self.Rset, (0, max_length - len(self.Rset)), constant_values=np.nan)
            Rreset_padded = np.pad(self.Rreset, (0, max_length - len(self.Rreset)), constant_values=np.nan)

            # Convert Vreset to a pandas DataFrame and save to Excel
            df = pd.DataFrame({'Vreset': Vreset_padded, 'Vset': Vset_padded, 'Rset' : Rset_padded, 'Rreset' : Rreset_padded})
            try:
                # Try to save the DataFrame to Excel
                df.to_excel(file_path, index=False)
                QMessageBox.information(self, 'Export Status', f"Exported: {len(self.Vset)} cycles!")

            except ValueError as e:
                if "This sheet is too large!" in str(e):
                    # Display a message box if the error is due to sheet size
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Warning)
                    msg.setText("Export failed")
                    msg.setInformativeText("The data size is too big to be exported to a single Excel sheet.")
                    msg.setWindowTitle("Export Error")
                    msg.exec_()
                else:
                    # Reraise any other exceptions
                    raise e

    ########################## Vset calculation for the current plotted Vset (Vreset stays N/A) #########################################
    def write_Vset(self, df):
        forth_idxes = np.where(np.gradient(df.Voltage)>0)[0]
        forth_idxes = forth_idxes[df.Voltage.iloc[forth_idxes] > 0.05]
        setvoltage = df.Voltage.iloc[forth_idxes]
        setcurrent = df.Current.iloc[forth_idxes]
        settime = df.TimeOutput.iloc[forth_idxes]
        R = setvoltage.abs()/setcurrent.abs()
        dR = -(np.gradient(np.log(R)))
        max_idx,_ = find_peaks(dR)
        
        highest_peak = max_idx[np.argmax(dR[max_idx])]
        threshold = 0.7 * dR[highest_peak]
        max_idx2 = max_idx[dR[max_idx] >= threshold]
        max_dR = dR[max_idx2]
        max_t = settime.iloc[max_idx2]

        if max_dR.size == 0:
            VSET = np.min(setvoltage)
        else:
            idx_RESET = np.where(settime == max_t.iloc[-1])[0][0]
            VSET = setvoltage.iloc[idx_RESET]
            if abs(VSET) < 0.05:
                VSET = setvoltage.iloc[-1]
        return VSET
    
    ########################## Vreset calculation for the current plotted Vreset (Vset stays N/A) #########################################
    def write_Vreset(self, df):
        # Take only the sweep in and not the sweep back part 
        # and also get rid of the first part of the sweep because it messes with the
        forth_idxes = np.where(np.gradient(df.Voltage)<0)[0]
        forth_idxes = forth_idxes[df.Voltage.iloc[forth_idxes] < -0.05]
        resvoltage = df.Voltage.iloc[forth_idxes]
        rescurrent = df.Current.iloc[forth_idxes]
        restime = df.TimeOutput.iloc[forth_idxes]
        R = resvoltage.abs()/rescurrent.abs()
        dR = np.gradient(np.log(R))
        max_idx,_ = find_peaks(dR)
        
        highest_peak = max_idx[np.argmax(dR[max_idx])]
        threshold = 0.7 * dR[highest_peak]
        max_idx2 = max_idx[dR[max_idx] >= threshold]
        
        max_dR = dR[max_idx2]
        max_t = restime.iloc[max_idx2]

        if max_dR.size == 0:
            VRESET = np.min(resvoltage)
        else:
            idx_RESET = np.where(restime == max_t.iloc[-1])[0][0]
            VRESET = resvoltage.iloc[idx_RESET]
            if abs(VRESET) < 0.05:
                VRESET = resvoltage.iloc[-1]
        return VRESET
    
    ##################### Same algorithm as write_Vset but calculates all Vsets of all cycles ###########################
    def update_Vset(self, dfset):
        if len(dfset) == 0:
            return
        for set_sweep in dfset:
            # Take only the sweep in and not the sweep back part 
            # and also get rid of the first part of the sweep because it messes with the gradient
            forth_idxes = np.where(np.gradient(set_sweep.Voltage)>0)[0]
            forth_idxes = forth_idxes[set_sweep.Voltage.iloc[forth_idxes] > 0.05]
            setvoltage = set_sweep.Voltage.iloc[forth_idxes]
            setcurrent = set_sweep.Current.iloc[forth_idxes]
            settime = set_sweep.TimeOutput.iloc[forth_idxes]
            R = setvoltage.abs()/setcurrent.abs()
            dR = -(np.gradient(np.log(R)))
            max_idx,_ = find_peaks(dR)
            
            highest_peak = max_idx[np.argmax(dR[max_idx])]
            threshold = 0.7 * dR[highest_peak]
            max_idx2 = max_idx[dR[max_idx] >= threshold]
            max_dR = dR[max_idx2]
            max_t = settime.iloc[max_idx2]

            if max_dR.size == 0:
                VSET = np.min(setvoltage)
            else:
                idx_RESET = np.where(settime == max_t.iloc[-1])[0][0]
                VSET = setvoltage.iloc[idx_RESET]
                if abs(VSET) < 0.05:
                    VSET = setvoltage.iloc[-1]
            self.Vset.append(VSET)

    ##################### Same algorithm as write_Vreset but calculates all Vresets of all cycles ###########################
    def update_Vreset(self, dfreset):
        if len(dfreset) == 0:
            return
        for reset_sweep in dfreset:
            # Take only the sweep in and not the sweep back part 
            # and also get rid of the first part of the sweep because it messes with the
            forth_idxes = np.where(np.gradient(reset_sweep.Voltage)<0)[0]
            forth_idxes = forth_idxes[reset_sweep.Voltage.iloc[forth_idxes] < -0.05]
            resvoltage = reset_sweep.Voltage.iloc[forth_idxes]
            rescurrent = reset_sweep.Current.iloc[forth_idxes]
            restime = reset_sweep.TimeOutput.iloc[forth_idxes]
            R = resvoltage.abs()/rescurrent.abs()
            dR = np.gradient(np.log(R))
            max_idx,_ = find_peaks(dR)
            
            highest_peak = max_idx[np.argmax(dR[max_idx])]
            threshold = 0.7 * dR[highest_peak]
            max_idx2 = max_idx[dR[max_idx] >= threshold]
            
            max_dR = dR[max_idx2]
            max_t = restime.iloc[max_idx2]

            if max_dR.size == 0:
                VRESET = np.min(resvoltage)
            else:
                idx_RESET = np.where(restime == max_t.iloc[-1])[0][0]
                VRESET = resvoltage.iloc[idx_RESET]
                if abs(VRESET) < 0.05:
                    VRESET = resvoltage.iloc[-1]
            self.Vreset.append(VRESET)
        
    ############################ Main Plot ##################################
    def plot_data(self, df):
        # Clear the previous plot
        self.figure.clear()

        # Create new axes
        self.ax1 = self.figure.add_subplot(111)

        # Left axis for Voltage Channel
        self.ax1.set_xlabel('Voltage')
        self.ax1.set_ylabel('Current', color=self.voltage_color)
        line1, = self.ax1.plot(df[self.VChannel], df[self.IChannel].abs(), color=self.voltage_color, linewidth=1.5, alpha=0.8, label=self.VChannel)
        #self.ax1.plot(df[self.VChannel], df[self.IChannel].abs(), color="blue", linewidth=0.5)
        self.ax1.tick_params(axis='y', labelcolor=self.voltage_color)
        if self.checkbox_grid.isChecked():
            self.ax1.grid(True, which='both', linestyle='--', linewidth=0.5)
        self.ax1.set_yscale('log')

        # Format y-axis with scientific notation
        self.ax1.yaxis.set_major_formatter(LogFormatterMathtext(base=10))

        # Add interactive cursor
        mplcursors.cursor([line1], hover=2)
        
        self.figure.tight_layout()
        self.canvas.draw()

    ####################### Voltage CDF Plot ###########################
    def plot_cdf_V(self):
        # Sort the data
        Vset_sorted = np.sort(self.Vset)
        Vreset_sorted = np.sort(self.Vreset)

        # Compute the cumulative probabilities
        cdf_Vset = np.arange(1, len(Vset_sorted) + 1) / len(Vset_sorted)
        cdf_Vreset = np.arange(1, len(Vreset_sorted) + 1) / len(Vreset_sorted)


        # Clear the previous plot
        self.figure2.clear()

        # Create new axes
        self.ax2 = self.figure2.add_subplot(111)

        # Left axis for Voltage Channel
        self.ax2.set_xlabel('Voltage')
        self.ax2.set_ylabel('Probability', color='black')
        line1, = self.ax2.plot(Vset_sorted, cdf_Vset, color=self.set_color, marker='.', linestyle='none', label='Vset')
        line2, = self.ax2.plot(Vreset_sorted, cdf_Vreset, color=self.reset_color, marker='.', linestyle='none', label='Vreset')
        if self.checkbox_grid.isChecked():
            self.ax2.grid(True)
        self.ax2.legend()
        
        # Add interactive cursor
        cursor1 = mplcursors.cursor([line1], hover=2)
        cursor2 = mplcursors.cursor([line2], hover=2)

        self.figure2.tight_layout()
        self.canvas2.draw()

    #################### Resistance CDF Plot #################################
    def plot_cdf_R(self):
        # Sort the data
        Rset_sorted = np.sort(self.Rset)
        Rreset_sorted = np.sort(self.Rreset)

        # Compute the cumulative probabilities
        cdf_Rset = np.arange(1, len(Rset_sorted) + 1) / len(Rset_sorted)
        cdf_Rreset = np.arange(1, len(Rreset_sorted) + 1) / len(Rreset_sorted)


        # Clear the previous plot
        self.figure3.clear()

        # Create new axes
        self.ax3 = self.figure3.add_subplot(111)

        # Left axis for Voltage Channel
        self.ax3.set_xlabel('Resistance (Ohm)')
        self.ax3.set_ylabel('Probability', color='black')
        line1, = self.ax3.plot(Rset_sorted, cdf_Rset, color=self.set_color, marker='.', linestyle='none', label='Rset')
        line2, = self.ax3.plot(Rreset_sorted, cdf_Rreset, color=self.reset_color, marker='.', linestyle='none', label='Rreset')
        if self.checkbox_grid.isChecked():
            self.ax3.grid(True)
        self.ax3.set_xscale('log')
        self.ax3.legend()
        
        # Add interactive cursor
        cursor1 = mplcursors.cursor([line1], hover=2)
        cursor2 = mplcursors.cursor([line2], hover=2)

        self.figure3.tight_layout()
        self.canvas3.draw()

    ###################### Resistance Plot ###############################
    def plot_res(self):
        #cycles = max(len(self.Rset), len(self.Rreset))

        # Clear the previous plot
        self.figure4.clear()
        
        x_lr = np.arange(1,len(self.Rset)+1)
        x_hr = np.arange(1,len(self.Rreset)+1)

        # Create new axes
        self.ax4 = self.figure4.add_subplot(111)

        # Left axis for Voltage Channel
        self.ax4.set_xlabel('Nr. of Cycle')
        self.ax4.set_ylabel('Resistance', color='blue')
        line1 = self.ax4.scatter(x_lr, self.Rset, color='red', label='Low Resistance')
        line2 = self.ax4.scatter(x_hr, self.Rreset, color='blue', label='High Resistance')
        if self.checkbox_grid.isChecked():
            self.ax4.grid(True)
        self.ax4.set_yscale('log')
        self.ax4.legend()
        
        # Add interactive cursor
        cursor1 = mplcursors.cursor(line1, hover=2)
        cursor2 = mplcursors.cursor(line2, hover=2)

        self.figure4.tight_layout()
        self.canvas4.draw()


    ####### Functions to add right click option to save plot ########
    #################################################################
    def show_context_menu(self, pos):
        context_menu = QMenu(self)
        save_action = context_menu.addAction("Save Plot")
        action = context_menu.exec_(self.canvas.mapToGlobal(pos))
        if action == save_action:
            self.save_plot(self.figure)

    def show_context_menu2(self, pos):
        context_menu = QMenu(self)
        save_action = context_menu.addAction("Save Plot")
        action = context_menu.exec_(self.canvas2.mapToGlobal(pos))
        if action == save_action:
            self.save_plot(self.figure2)

    def show_context_menu3(self, pos):
        context_menu = QMenu(self)
        save_action = context_menu.addAction("Save Plot")
        action = context_menu.exec_(self.canvas3.mapToGlobal(pos))
        if action == save_action:
            self.save_plot(self.figure3)

    def show_context_menu4(self, pos):
        context_menu = QMenu(self)
        save_action = context_menu.addAction("Save Plot")
        action = context_menu.exec_(self.canvas4.mapToGlobal(pos))
        if action == save_action:
            self.save_plot(self.figure4)

    def save_plot(self, figure):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Plot", "", "PNG Files (*.png);;All Files (*)", options=options)
        if file_path:
            figure.savefig(file_path)

    def pick_voltage_color(self):
        # Open color picker dialog and set the selected color
        color = QColorDialog.getColor()
        if color.isValid():
            self.voltage_color = color.name()  # Convert selected color to HEX code (e.g., "#RRGGBB")
            self.voltage_color_button.setStyleSheet(f"background-color: {self.voltage_color};")  # Update button background color
        self.update_plot()

    def pick_set_color(self):
        # Open color picker dialog and set the selected color
        color = QColorDialog.getColor()
        if color.isValid():
            self.set_color = color.name()  # Convert selected color to HEX code (e.g., "#RRGGBB")
            self.set_color_button.setStyleSheet(f"background-color: {self.set_color};")  # Update button background color
        self.update_plot()

    def pick_reset_color(self):
        # Open color picker dialog and set the selected color
        color = QColorDialog.getColor()
        if color.isValid():
            self.reset_color = color.name()  # Convert selected color to HEX code (e.g., "#RRGGBB")
            self.reset_color_button.setStyleSheet(f"background-color: {self.reset_color};")  # Update button background color
        self.update_plot()